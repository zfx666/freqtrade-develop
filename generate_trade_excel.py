"""根据策略信号模拟交易并生成Excel"""
import pandas as pd
import numpy as np
from pathlib import Path

print("=" * 80)
print("模拟交易并生成Excel分析")
print("=" * 80)

# 读取策略CSV
csv_file = Path('no_trades_analysis.csv')
if not csv_file.exists():
    print("❌ 未找到 no_trades_analysis.csv，请先运行 analyze_trades.py")
    exit(1)

df = pd.read_csv(csv_file, encoding='utf-8-sig')
print(f"\n读取数据: {len(df)} 行")

# 转换时间
df['时间'] = pd.to_datetime(df['时间'])

# 找到所有入场信号
entry_signals = df[df['入场信号'] == True].copy()
print(f"入场信号: {len(entry_signals)} 个")

if len(entry_signals) == 0:
    print("❌ 没有入场信号！")
    exit(1)

# 模拟交易
trades = []
stop_loss_pct = 0.01  # 1%止损

for idx, entry_row in entry_signals.iterrows():
    entry_time = entry_row['时间']
    entry_price = entry_row['收盘价']
    stop_loss_price = entry_price * (1 - stop_loss_pct)
    
    # 找到入场后的数据
    future_data = df[df['时间'] > entry_time].copy()
    
    if len(future_data) == 0:
        continue  # 数据末尾，无法模拟
    
    # 寻找出场点
    exit_time = None
    exit_price = None
    exit_reason = None
    exit_detail = None
    
    for future_idx, row in future_data.iterrows():
        current_price = row['收盘价']
        current_time = row['时间']
        
        # 检查出场条件
        
        # 条件1：1%止损
        if current_price <= stop_loss_price:
            exit_time = current_time
            exit_price = current_price
            exit_reason = "止损"
            exit_detail = f"价格{current_price:.2f}触发1%止损线{stop_loss_price:.2f}"
            break
        
        # 条件2：结构转弱（当前低点 < 前一低点）
        prev_idx = future_data.index.get_loc(future_idx) - 1
        if prev_idx >= 0:
            prev_row_idx = future_data.index[prev_idx]
            prev_low = df.loc[prev_row_idx, '最低价']
            curr_low = row['最低价']
            
            if curr_low < prev_low:
                exit_time = current_time
                exit_price = current_price
                exit_reason = "结构转弱"
                exit_detail = f"低点{curr_low:.2f} < 前低点{prev_low:.2f}"
                break
        
        # 条件3：4h失势（收盘价 <= 下轨）
        if pd.notna(row.get('跌破下轨4h')) and row['跌破下轨4h']:
            exit_time = current_time
            exit_price = current_price
            exit_reason = "4h失势"
            bb_lower = row.get('布林下轨4h', 0)
            exit_detail = f"收盘{current_price:.2f} <= 下轨{bb_lower:.2f}"
            break
    
    # 如果没有触发出场，使用最后一根K线
    if exit_time is None:
        last_row = future_data.iloc[-1]
        exit_time = last_row['时间']
        exit_price = last_row['收盘价']
        exit_reason = "数据结束"
        exit_detail = "回测时间结束"
    
    # 计算盈亏
    profit_pct = (exit_price - entry_price) / entry_price * 100
    profit_abs = (exit_price - entry_price)  # 假设每次1个单位
    
    # 持仓时长
    duration_hours = (exit_time - entry_time).total_seconds() / 3600
    
    # 获取出场时的数据
    exit_row = df[df['时间'] == exit_time].iloc[0] if exit_time else future_data.iloc[-1]
    
    trade = {
        '交易序号': len(trades) + 1,
        # === 入场信息 ===
        '入场时间': entry_time,
        '入场_开盘价': entry_row.get('开盘价', 0),
        '入场_最高价': entry_row.get('最高价', 0),
        '入场_最低价': entry_row.get('最低价', 0),
        '入场_收盘价': entry_price,
        '入场_成交量': entry_row.get('成交量', 0),
        
        # === 入场时4h布林带参数 ===
        '入场_4h收盘价': entry_row.get('4小时收盘价', 0),
        '入场_4h上轨': entry_row.get('布林上轨4h', 0),
        '入场_4h中轨': entry_row.get('布林中轨4h', 0),
        '入场_4h下轨': entry_row.get('布林下轨4h', 0),
        '入场_4h宽度%': entry_row.get('布林宽度4h', 0) * 100,
        '入场_价格vs上轨': entry_row.get('4小时收盘价', 0) - entry_row.get('布林上轨4h', 0),
        
        # === 入场时1h布林带 ===
        '入场_1h上轨': entry_row.get('布林上轨1h', 0),
        '入场_1h中轨': entry_row.get('布林中轨1h', 0),
        '入场_1h下轨': entry_row.get('布林下轨1h', 0),
        
        # === 入场条件 ===
        '入场_宽度达标': '是' if entry_row.get('宽度达标4h') else '否',
        '入场_上穿上轨': '是' if entry_row.get('上穿上轨4h') else '否',
        '入场_Armed当根': '是' if entry_row.get('Armed当根4h') else '否',
        '入场_Armed持续': '是' if entry_row.get('Armed持续4h') else '否',
        '入场_HLH信号': '是' if entry_row.get('HLH信号') else '否',
        
        # === 入场时结构信息 ===
        '入场_结构高': entry_row.get('结构高', 0),
        '入场_结构低': entry_row.get('结构低', 0),
        '入场_结构类型': entry_row.get('结构类型', 0),
        
        # === 交易执行 ===
        '入场信号': 1,  # 用1表示
        '止损价(1%)': stop_loss_price,
        '止损距离%': stop_loss_pct * 100,
        
        # === 出场信息 ===
        '出场时间': exit_time,
        '出场_收盘价': exit_price,
        '出场_4h收盘价': exit_row.get('4小时收盘价', 0),
        '出场_4h下轨': exit_row.get('布林下轨4h', 0),
        '出场_4h宽度%': exit_row.get('布林宽度4h', 0) * 100,
        '出场_Armed状态': '是' if exit_row.get('Armed持续4h') else '否',
        '出场_跌破下轨': '是' if exit_row.get('跌破下轨4h') else '否',
        
        '出场原因': exit_reason,
        '出场详情': exit_detail,
        
        # === 交易结果 ===
        '持仓时长(小时)': duration_hours,
        '盈亏点数': exit_price - entry_price,
        '盈亏%': profit_pct,
        '盈亏金额': profit_abs,
        '盈亏状态': '盈利' if profit_pct > 0 else '亏损',
        
        # === 最大回撤（持仓期间）===
        '持仓期最高价': future_data.loc[:future_data[future_data['时间']==exit_time].index[0], '最高价'].max() if exit_time else 0,
        '持仓期最低价': future_data.loc[:future_data[future_data['时间']==exit_time].index[0], '最低价'].min() if exit_time else 0,
    }
    
    trades.append(trade)

trades_df = pd.DataFrame(trades)

# 添加累计盈亏
trades_df['累计盈亏%'] = trades_df['盈亏%'].cumsum()
trades_df['累计盈亏金额'] = trades_df['盈亏金额'].cumsum()

print(f"\n模拟交易数: {len(trades_df)}")
print(f"总盈亏: {trades_df['盈亏金额'].sum():.2f}")
print(f"胜率: {len(trades_df[trades_df['盈亏%']>0])}/{len(trades_df)} = {len(trades_df[trades_df['盈亏%']>0])/len(trades_df)*100:.1f}%")

# 导出Excel（带格式）
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "交易明细"
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(trades_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 表头加粗
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 入场信号列标红
                if c_idx == trades_df.columns.get_loc('入场信号') + 1:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    cell.font = Font(color="FF0000", bold=True)
                
                # 盈亏状态上色
                if c_idx == trades_df.columns.get_loc('盈亏状态') + 1:
                    if value == '盈利':
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        cell.font = Font(color="008000", bold=True)
                    elif value == '亏损':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        cell.font = Font(color="FF0000", bold=True)
                
                # 盈亏%上色
                if c_idx == trades_df.columns.get_loc('盈亏%') + 1:
                    try:
                        if float(value) > 0:
                            cell.font = Font(color="008000")
                        elif float(value) < 0:
                            cell.font = Font(color="FF0000")
                    except:
                        pass
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    excel_file = 'trade_analysis_detailed.xlsx'
    wb.save(excel_file)
    print(f"\n✅ Excel文件已生成: {excel_file}")
    print("   - 入场信号列已标红")
    print("   - 盈亏状态已上色（绿色=盈利，红色=亏损）")
    
except ImportError:
    # 如果没有openpyxl，导出CSV
    csv_file = 'trade_analysis_detailed.csv'
    trades_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"\n✅ CSV文件已生成: {csv_file}")
    print("   （需要安装openpyxl才能生成带格式的Excel文件）")

# 统计分析
print("\n" + "=" * 80)
print("交易统计分析")
print("=" * 80)

wins = trades_df[trades_df['盈亏%'] > 0]
losses = trades_df[trades_df['盈亏%'] <= 0]

print(f"\n盈利交易: {len(wins)} 笔")
print(f"  平均盈利: {wins['盈亏%'].mean():.2f}%")
print(f"  最大盈利: {wins['盈亏%'].max():.2f}%")

print(f"\n亏损交易: {len(losses)} 笔")
print(f"  平均亏损: {losses['盈亏%'].mean():.2f}%")
print(f"  最大亏损: {losses['盈亏%'].min():.2f}%")

print(f"\n出场原因统计:")
for reason, count in trades_df['出场原因'].value_counts().items():
    pct = count / len(trades_df) * 100
    print(f"  {reason}: {count} 笔 ({pct:.1f}%)")

print(f"\n前10笔交易:")
display_cols = ['交易序号', '入场时间', '入场_收盘价', '出场时间', '出场_收盘价', '盈亏%', '出场原因']
print(trades_df[display_cols].head(10).to_string(index=False))

print("\n" + "=" * 80)

